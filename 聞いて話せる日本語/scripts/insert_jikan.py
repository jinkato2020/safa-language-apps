# 時の表現シートを 文法3.xlsx の23番目に挿入し、全シートの番号(NN_)を振り直して上書き保存。
import json, re, os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
P = os.path.join(ROOT, "文法3.xlsx")
items = json.load(open(os.path.join(ROOT, "scripts", "_jikan.json"), encoding="utf-8"))
NEW_REST = "時の表現「〜とき・〜まえに・〜あとで」"
INSERT_AT = 22  # 0-based → 23番目

wb = load_workbook(P)
ws = wb.create_sheet("__jikan_new__")
hdr = Font(bold=True)
ws["A1"] = "日本語"; ws["B1"] = "ネパール語"; ws["A1"].font = hdr; ws["B1"].font = hdr
for i, it in enumerate(items[:20]):
    ws.cell(row=i+2, column=1, value=it["jp"])
    ws.cell(row=i+2, column=2, value=it["ne"])
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 48
for row in ws.iter_rows(min_row=2, max_row=21, max_col=2):
    for c in row:
        c.alignment = Alignment(vertical="center", wrap_text=True)

# 並べ替え: 新シートを INSERT_AT へ
wb._sheets.remove(ws)
wb._sheets.insert(INSERT_AT, ws)

# 番号振り直し (衝突回避のため一旦tmp名→最終名)
def rest(name): return re.sub(r'^\d+_', '', name)
def sanitize(n): return re.sub(r'[\\/?*\[\]:]', '・', n)[:31]
order = list(wb._sheets)
rest_orig = [rest(s.title) for s in order]  # 旧テーマ名(番号除く)を先に保持
for i, s in enumerate(order):
    s.title = f"_tmp{i}_"
for i, s in enumerate(order):
    r = NEW_REST if s is ws else rest_orig[i]
    s.title = sanitize(f"{i+1:02d}_{r}")

wb.save(P)
print(f"保存: {P}")
for i, s in enumerate(wb.sheetnames):
    print(f"  [{i+1:02d}] {s}")
