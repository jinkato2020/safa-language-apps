# 文法.xlsx の新カリキュラムを App B に適用(日本語)。
# grammarExamples.json を置換(jp+ne, 各20)、grammarThemes.json と i18n ja のテーマ名更新。
# 旧 grammarExamples.json は .bak に退避。ベンガル語訳/音声/ne・bnテーマ名は後工程。
import json, re, os, shutil
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "文法.xlsx")
DATA = os.path.join(BASE, "expo-app", "data")
GE = os.path.join(DATA, "grammarExamples.json")
GT = os.path.join(DATA, "grammarThemes.json")
JA = os.path.join(BASE, "expo-app", "src", "i18n", "ja.json")

wb = load_workbook(XLSX, read_only=True)
grammar = {}; names = {}
for idx, sn in enumerate(wb.sheetnames, start=1):
    name = re.sub(r'^\d+_', '', sn)
    ws = wb[sn]
    items = []
    for r in range(2, 22):  # 例題1-20
        jp = ws.cell(r, 1).value
        ne = ws.cell(r, 2).value
        if jp:
            items.append({"jp": str(jp).strip(), "ne": (str(ne).strip() if ne else "")})
    grammar[str(idx)] = items
    names[str(idx)] = name
wb.close()

bad = [k for k, v in grammar.items() if len(v) != 20]
assert not bad, f"20文でないテーマ: {bad}"

shutil.copy2(GE, GE + ".bak")
json.dump(grammar, open(GE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

gt = json.load(open(GT, encoding="utf-8"))
for t in gt:
    t["name"] = names[str(t["id"])]
    t["exampleCount"] = 20
json.dump(gt, open(GT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

ja = json.load(open(JA, encoding="utf-8"))
ja["grammarThemes"] = names
json.dump(ja, open(JA, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

print("適用完了:")
print(f"  grammarExamples.json: {len(grammar)}テーマ x 20文 (旧は .bak)")
print("  grammarThemes.json / i18n ja: 名称更新")
for k in ["1", "23", "29", "30"]:
    print(f"   {k}: {names[k]}")
