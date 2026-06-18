# 中国語 TTS→STT差分＋声調確認。zh-CNロケールでSTT→ピンイン+声調(TONE3)で参照と照合。
#  同音字・繁簡は同ピンインでPASS。声調違い=声調エラー、ピンイン違い=発音エラーとして抽出。
#  対象: AppB_着色全リストのzh134件(char不一致=候補)。音声=audio/zh/{conv,gram}/<id>.mp3。
import os, re, json, subprocess, tempfile, urllib.request, urllib.error, sys
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import Font, PatternFill
from pypinyin import pinyin, Style
sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env = {}
for l in open(os.path.join(ROOT, ".env"), encoding="utf-8"):
    m = re.match(r"\s*([A-Za-z0-9_]+)\s*=\s*(.*)\s*$", l)
    if m: env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
KEY, REG = env["AZURE_SPEECH_KEY"], env["AZURE_SPEECH_REGION"]
EP = f"https://{REG}.stt.speech.microsoft.com/speech/recognition/conversation/cognitiveservices/v1?language=zh-CN&format=detailed"
NUMH = set("〇零一二三四五六七八九十百千万亿億兩两")
cjk = lambda s: "".join(c for c in str(s or "") if 0x4E00 <= ord(c) <= 0x9FFF)
def py(s, tone=True):
    st = Style.TONE3 if tone else Style.NORMAL
    return [x[0] for x in pinyin(s, style=st, errors="ignore")]
def stripnum(s): return "".join(c for c in s if c not in NUMH)

wb = openpyxl.load_workbook(os.path.join(ROOT, "AppB_着色全リスト.xlsx")); ws = wb.active
hdr = [c.value for c in ws[1]]; Li, Ni, Ti = hdr.index("言語"), hdr.index("番号"), hdr.index("テキスト")
items = [(str(ws.cell(r, Ni+1).value), str(ws.cell(r, Ti+1).value))
         for r in range(2, ws.max_row+1) if str(ws.cell(r, Li+1).value) == "zh"]
print(f"zh対象: {len(items)}件", flush=True)

def chk(idv, text):
    a, b, c = (idv.split("-") + [None, None])[:3]
    mode = "conv" if c is not None else "gram"
    mp3 = os.path.join(ROOT, "audio", "zh", mode, idv + ".mp3")
    if not os.path.exists(mp3): return (idv, text, "", "NOAUDIO")
    wav = tempfile.mktemp(suffix=".wav")
    try:
        subprocess.run(["ffmpeg","-y","-loglevel","error","-i",mp3,"-ar","16000","-ac","1",wav],check=True,capture_output=True)
        hd = {"Ocp-Apim-Subscription-Key": KEY, "Content-Type": "audio/wav; codecs=audio/pcm; samplerate=16000", "Accept": "application/json"}
        rec = ""
        for at in range(4):
            try:
                r = urllib.request.urlopen(urllib.request.Request(EP, data=open(wav,"rb").read(), headers=hd, method="POST"), timeout=60)
                nb = (json.loads(r.read().decode()).get("NBest") or [{}]); rec = nb[0].get("Display","") if nb else ""; break
            except urllib.error.HTTPError as e:
                if e.code in (429,500,502,503) and at<3: continue
                return (idv, text, f"HTTP{e.code}", "ERR")
            except Exception:
                if at<3: continue
                return (idv, text, "timeout", "ERR")
        rt, rc = cjk(text), cjk(rec)
        ref_t, rec_t = py(rt, True), py(rc, True)            # ピンイン+声調
        if ref_t == rec_t: return (idv, text, rec, "アーティファクト(同音/繁簡)")
        if py(stripnum(rt), True) == py(stripnum(rc), True): return (idv, text, rec, "数字表記差")
        ref_b, rec_b = py(rt, False), py(rc, False)          # ピンイン(声調無視)
        if ref_b == rec_b: return (idv, text, rec, "★声調エラー候補")
        return (idv, text, rec, "発音エラー候補")
    finally:
        if os.path.exists(wav):
            try: os.remove(wav)
            except OSError: pass

res = list(ThreadPoolExecutor(max_workers=6).map(lambda it: chk(*it), items))
import collections
cnt = collections.Counter(r[3] for r in res)
print("\n=== 判定集計 ===")
for k, v in cnt.most_common(): print(f"  {k}: {v}")
print("\n=== 声調/発音エラー候補 ===")
for idv, text, rec, verdict in res:
    if "エラー候補" in verdict:
        print(f"  [{verdict}] {idv}: 本文[{text[:24]}] 認識[{rec[:24]}]")
        print(f"      本文py={py(cjk(text))}\n      認識py={py(cjk(rec))}")
# Excel
out = os.path.join(ROOT, "AppB_中国語_声調確認.xlsx")
w2 = openpyxl.Workbook(); s = w2.active; s.title = "zh声調確認"
s.append(["番号","判定","本文","認識(zh-CN)","本文ピンイン","認識ピンイン"])
for c in s[1]: c.font = Font(bold=True)
order = {"発音エラー候補":0,"★声調エラー候補":1,"数字表記差":2,"アーティファクト(同音/繁簡)":3,"NOAUDIO":4,"ERR":4}
for idv, text, rec, verdict in sorted(res, key=lambda r: order.get(r[3],9)):
    s.append([idv, verdict, text, rec, " ".join(py(cjk(text))), " ".join(py(cjk(rec)))])
    if "エラー候補" in verdict:
        for c in range(1,7): s.cell(s.max_row,c).fill = PatternFill("solid", fgColor="FFC7CE")
for col,wd in zip("ABCDEF",(10,22,30,30,34,34)): s.column_dimensions[col].width = wd
w2.save(out)
print("\n出力:", os.path.basename(out))
