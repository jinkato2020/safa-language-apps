# App B 日本語音声 全2400件 読み検査(Azure発音評価のかな認識 × jp-reading 正解かな 差分)。
#  参照かなを渡すとAzureが「実際に鳴った音」をかな認識→正規化差分で誤読を検出。
#  会話 audio/ja/conv (examples.json) + 文法 audio/ja/gram (grammarExamples.json)。
#  再開可能(_pa_scan_results.json)。出力: AppB_読み要確認.xlsx。
import os, re, json, base64, subprocess, tempfile, urllib.request, urllib.error, sys, time, difflib
from concurrent.futures import ThreadPoolExecutor
sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
APP_ROOT = os.path.dirname(ROOT)
env = {}
for l in open(os.path.join(ROOT, ".env"), encoding="utf-8"):
    m = re.match(r"\s*([A-Za-z0-9_]+)\s*=\s*(.*)\s*$", l)
    if m: env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
KEY, REG = env["AZURE_SPEECH_KEY"], env["AZURE_SPEECH_REGION"]
EP = f"https://{REG}.stt.speech.microsoft.com/speech/recognition/conversation/cognitiveservices/v1?language=ja-JP&format=detailed"
ex = json.load(open(os.path.join(ROOT, "expo-app/data/examples.json"), encoding="utf-8"))
gr = json.load(open(os.path.join(ROOT, "expo-app/data/grammarExamples.json"), encoding="utf-8"))
jr = json.load(open(os.path.join(ROOT, "expo-app/data/jp-reading.json"), encoding="utf-8"))
RES_PATH = os.path.join(ROOT, "scripts/_pa_scan_results.json")
XLSX = os.path.join(ROOT, "AppB_読み要確認.xlsx")
norm = lambda s: re.sub(r"[、。？！\s]", "", s or "")

items = []
for k, arr in ex.items():
    for i, e in enumerate(arr): items.append((f"{k}-{i+1}", "conv", e.get("jp")))
for k, arr in gr.items():
    for i, e in enumerate(arr): items.append((f"{k}-{i+1}", "gram", e.get("jp") if isinstance(e, dict) else e))

results = {}
if os.path.exists(RES_PATH):
    try: results = json.load(open(RES_PATH, encoding="utf-8"))
    except Exception: results = {}
todo = [it for it in items if it[0] not in results]
print(f"全{len(items)}件 / 残{len(todo)}件 を検査", flush=True)

def assess(idv, mode, jp):
    wav = None
    try:
        ref = jr.get(jp, {}).get("kana", "")
        mp3 = os.path.join(ROOT, "audio", "ja", mode, idv + ".mp3")
        wav = tempfile.mktemp(suffix=".wav")
        subprocess.run(["ffmpeg", "-y", "-loglevel", "error", "-i", mp3, "-ar", "16000", "-ac", "1", wav],
                       check=True, capture_output=True)
        cfg = {"ReferenceText": ref.replace("、", " ").replace("。", "").replace("？", "").strip(),
               "GradingSystem": "HundredMark", "Granularity": "Word", "EnableMiscue": True}
        hdr = {"Ocp-Apim-Subscription-Key": KEY, "Content-Type": "audio/wav; codecs=audio/pcm; samplerate=16000",
               "Accept": "application/json", "Pronunciation-Assessment": base64.b64encode(json.dumps(cfg).encode()).decode()}
        data = open(wav, "rb").read()
        rec = ""
        for a in range(5):
            try:
                r = urllib.request.urlopen(urllib.request.Request(EP, data=data, headers=hdr, method="POST"), timeout=60)
                res = json.loads(r.read().decode()); nb = (res.get("NBest") or [{}])[0]
                rec = nb.get("Display") or nb.get("Lexical") or ""
                break
            except urllib.error.HTTPError as e:
                if e.code in (429, 500, 502, 503) and a < 4: time.sleep(2 * (a + 1)); continue
                raise
            except Exception:
                if a < 4: time.sleep(2 * (a + 1)); continue
                raise
        R, G = norm(ref), norm(rec)
        ok = (R == G)
        diffs = []
        if not ok:
            for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(None, R, G).get_opcodes():
                if tag != "equal": diffs.append(f"参照[{R[i1:i2]}]↔音声[{G[j1:j2]}]")
        return idv, {"mode": mode, "jp": jp, "ref": ref, "rec": rec, "ok": ok, "diff": " / ".join(diffs)}
    except Exception as e:
        return idv, {"mode": mode, "jp": jp, "ref": jr.get(jp, {}).get("kana", ""), "rec": "", "ok": None, "diff": f"ERROR:{e}"}
    finally:
        if wav and os.path.exists(wav):
            try: os.remove(wav)
            except OSError: pass

done = 0
with ThreadPoolExecutor(max_workers=8) as exr:
    for idv, rec in exr.map(lambda it: assess(*it), todo):
        results[idv] = rec; done += 1
        if done % 50 == 0:
            json.dump(results, open(RES_PATH, "w", encoding="utf-8"), ensure_ascii=False)
            ng = sum(1 for v in results.values() if v["ok"] is False)
            print(f"  {done}/{len(todo)}  (現時点NG {ng})", flush=True)
json.dump(results, open(RES_PATH, "w", encoding="utf-8"), ensure_ascii=False)

# Excel 出力(要確認のみ)
import openpyxl
from openpyxl.styles import Font, PatternFill
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "読み要確認"
ws.append(["番号", "モード", "原文", "参照かな(正解)", "音声認識", "相違スパン"])
for c in ws[1]: c.font = Font(bold=True)
def sortkey(idv):
    p = idv.split("-"); return tuple(int(x) if x.isdigit() else 0 for x in p)
ng = [idv for idv in results if results[idv]["ok"] is False]
err = [idv for idv in results if results[idv]["ok"] is None]
for idv in sorted(ng + err, key=sortkey):
    v = results[idv]
    ws.append([idv, v["mode"], v["jp"], v["ref"], v["rec"], v["diff"]])
for col, w in zip("ABCDEF", (10, 7, 40, 34, 34, 40)): ws.column_dimensions[col].width = w
wb.save(XLSX)
total = len(results); okc = sum(1 for v in results.values() if v["ok"] is True)
print(f"\nDONE 全{total}  一致{okc}  要確認{len(ng)}  エラー{len(err)}  → {os.path.basename(XLSX)}")
