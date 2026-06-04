# _jp_grammar.json から 文法2.xlsx を作成。30シート(テーマごと)、A列=日本語 B列=ネパール語、各20文。
import json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 聞いて話せる日本語
SRC = os.path.join(ROOT, "scripts", "_jp_grammar.json")
OUT = os.path.join(ROOT, "文法2.xlsx")

data = json.load(open(SRC, encoding="utf-8"))
# 番号順
themes = sorted(data.values(), key=lambda t: int(t["num"]))

def sheet_name(num, name):
    # Excel禁止文字を除去、31字以内
    s = f"{int(num):02d}_{name}"
    s = re.sub(r'[\\/?*\[\]:]', '・', s)
    return s[:31]

wb = Workbook()
wb.remove(wb.active)
hdr = Font(bold=True)
for t in themes:
    ws = wb.create_sheet(sheet_name(t["num"], t["name"]))
    ws["A1"] = "日本語"; ws["B1"] = "ネパール語"
    ws["A1"].font = hdr; ws["B1"].font = hdr
    for i, it in enumerate(t["items"][:20]):
        ws.cell(row=i+2, column=1, value=it.get("jp", ""))
        ws.cell(row=i+2, column=2, value=it.get("ne", ""))
    # 例題21-40: 日本語(A列, 行22-41)。21-30はネパール語(B列)も追加、31-40はA列のみ。
    ne2 = t.get("ne2", [])
    for i, jp in enumerate(t.get("items2", [])[:20]):
        ws.cell(row=i+22, column=1, value=jp if isinstance(jp, str) else jp.get("jp", ""))
        if i < len(ne2):
            ws.cell(row=i+22, column=2, value=ne2[i])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 48
    for row in ws.iter_rows(min_row=2, max_row=41, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)

wb.save(OUT)
print(f"saved {OUT}: {len(themes)} sheets")
for t in themes:
    print(f"  {sheet_name(t['num'], t['name'])}: {len(t['items'])} 文")
