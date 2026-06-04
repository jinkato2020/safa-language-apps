# App B 会話を修正(ネパール言及→日本語学習向け)し examples.json に適用、
# 会話2.xlsx に出力。変更セルは緑で色付け。30シート、A列=日本語 B列=ネパール語。
import json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "expo-app", "data")
EXP = os.path.join(DATA, "examples.json")
ex = json.load(open(EXP, encoding="utf-8"))
themes = json.load(open(os.path.join(DATA, "themes.json"), encoding="utf-8"))
OUT = os.path.join(ROOT, "会話2.xlsx")

# 修正: key "T-L" -> { index(1-based): (jp, ne) }
FIXES = {
 "1-2": {10: ("日本の文化にとても興味があります。", "मलाई जापानी संस्कृतिमा धेरै रुचि छ।"),
         15: ("英語とスペイン語が話せますが、今は日本語を学んでいます。", "म अंग्रेजी र स्पेनिश बोल्न सक्छु, तर अहिले जापानी भाषा सिक्दैछु।")},
 "5-2": {20: ("日本料理を初めて食べましたが、とても気に入りました。", "पहिलो पटक जापानी खाना खाएँ, तर धेरै मन पर्‍यो।")},
 "19-2": {14: ("牛は昔から人間の生活に役立ってきた動物です。", "गाई पुरानो समयदेखि मानिसको जीवनमा उपयोगी जनावर हो।")},
 "25-2": {15: ("通訳サービスがあるか役所に確認してください。", "दोभाषे सेवा छ कि छैन भनी नगरपालिकामा सोध्नुहोस्।")},
}
# 適用 + 変更位置 (theme, level, index) を記録
modified = set()
for key, fixes in FIXES.items():
    t, l = map(int, key.split("-"))
    for idx, (jp, ne) in fixes.items():
        ex[key][idx-1] = {"jp": jp, "ne": ne}
        modified.add((t, l, idx))
json.dump(ex, open(EXP, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def sheet_name(num, name):
    return re.sub(r'[\\/?*\[\]:]', '・', f"{int(num):02d}_{name}")[:31]

wb = Workbook(); wb.remove(wb.active)
hdr = Font(bold=True)
lvlfill = PatternFill("solid", fgColor="FFF2CC")   # レベル見出し(黄)
modfill = PatternFill("solid", fgColor="C6EFCE")   # 変更セル(緑)
LV = {1: "初級", 2: "中級", 3: "上級"}

for t in themes:
    tid = t["id"]
    ws = wb.create_sheet(sheet_name(tid, t["name"]))
    ws["A1"] = "日本語"; ws["B1"] = "ネパール語"; ws["A1"].font = hdr; ws["B1"].font = hdr
    r = 2
    for lv in (1, 2, 3):
        arr = ex.get(f"{tid}-{lv}", [])
        c = ws.cell(row=r, column=1, value=f"── {LV[lv]} ──"); c.font = hdr; c.fill = lvlfill
        ws.cell(row=r, column=2).fill = lvlfill
        r += 1
        for i, it in enumerate(arr):
            ca = ws.cell(row=r, column=1, value=it.get("jp", ""))
            cb = ws.cell(row=r, column=2, value=it.get("ne", ""))
            if (tid, lv, i+1) in modified:
                ca.fill = modfill; cb.fill = modfill
            r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 50
    for row in ws.iter_rows(min_row=2, max_row=r, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

wb.save(OUT)
print(f"saved {OUT}: {len(wb.sheetnames)} sheets / 変更 {len(modified)}セル行")
for m in sorted(modified): print("  colored:", m)
