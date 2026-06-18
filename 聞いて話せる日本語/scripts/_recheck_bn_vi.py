# flagged bn/vi を「ロケール固定」Azure STT で再認識し参照テキストと照合。
#  言語自動判定の誤り(bn→ラテン/他文字, vi→中/タイ等)を排除して、真の不一致だけ残す。
import os, re, json, base64, subprocess, tempfile, urllib.request, urllib.error, sys, unicodedata
from concurrent.futures import ThreadPoolExecutor
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env = {}
for l in open(os.path.join(ROOT, ".env"), encoding="utf-8"):
    m = re.match(r"\s*([A-Za-z0-9_]+)\s*=\s*(.*)\s*$", l)
    if m: env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
KEY, REG = env["AZURE_SPEECH_KEY"], env["AZURE_SPEECH_REGION"]
LOCALE = {"bn": "bn-IN", "vi": "vi-VN"}
def ep(loc): return f"https://{REG}.stt.speech.microsoft.com/speech/recognition/conversation/cognitiveservices/v1?language={loc}&format=detailed"
NUM = set("0123456789")
def norm(s):
    s = unicodedata.normalize("NFC", str(s or "")).lower()
    return "".join(ch for ch in s if ch.isalnum())          # 記号/空白除去
def strip_num(s): return "".join(ch for ch in s if ch not in NUM)

wb = openpyxl.load_workbook(os.path.join(ROOT, "AppB_着色全リスト.xlsx")); ws = wb.active
hdr = [c.value for c in ws[1]]
Li, Ni, Ti = hdr.index("言語"), hdr.index("番号"), hdr.index("テキスト")
items = []
for r in range(2, ws.max_row + 1):
    lang = str(ws.cell(r, Li + 1).value)
    if lang in ("bn", "vi"):
        items.append((lang, str(ws.cell(r, Ni + 1).value), str(ws.cell(r, Ti + 1).value)))
print(f"対象: bn {sum(1 for x in items if x[0]=='bn')} / vi {sum(1 for x in items if x[0]=='vi')}", flush=True)

def stt(lang, idv, text):
    a, b, c = (idv.split("-") + [None, None])[:3]
    mode = "conv" if c is not None else "gram"
    mp3 = os.path.join(ROOT, "audio", lang, mode, idv + ".mp3")
    if not os.path.exists(mp3):
        return (lang, idv, text, "(no audio)", "NOAUDIO")
    wav = tempfile.mktemp(suffix=".wav")
    try:
        subprocess.run(["ffmpeg","-y","-loglevel","error","-i",mp3,"-ar","16000","-ac","1",wav],check=True,capture_output=True)
        hdrs = {"Ocp-Apim-Subscription-Key": KEY, "Content-Type": "audio/wav; codecs=audio/pcm; samplerate=16000", "Accept":"application/json"}
        data = open(wav, "rb").read()
        rec = ""
        for at in range(4):
            try:
                rr = urllib.request.urlopen(urllib.request.Request(ep(LOCALE[lang]), data=data, headers=hdrs, method="POST"), timeout=60)
                res = json.loads(rr.read().decode()); nb = (res.get("NBest") or [{}])
                rec = (nb[0].get("Display") or nb[0].get("Lexical") or "") if nb else ""
                break
            except urllib.error.HTTPError as e:
                if e.code in (429,500,502,503) and at < 3: continue
                return (lang, idv, text, f"HTTP{e.code}", "ERR")
            except Exception:
                if at < 3: continue
                return (lang, idv, text, "timeout", "ERR")
        t, g = norm(text), norm(rec)
        if t == g: verdict = "一致(アーティファクト確定)"
        elif strip_num(t) == strip_num(g): verdict = "数字表記差(音声OK)"
        else: verdict = "候補"
        return (lang, idv, text, rec, verdict)
    finally:
        if os.path.exists(wav):
            try: os.remove(wav)
            except OSError: pass

res = []
with ThreadPoolExecutor(max_workers=6) as ex:
    res = list(ex.map(lambda it: stt(*it), items))
import collections
by = collections.Counter((r[0], r[4]) for r in res)
print("\n=== 判定集計 (言語, 判定): 件数 ===")
for k, v in sorted(by.items()): print(f"  {k}: {v}")
print("\n=== 候補(真の不一致の可能性) ===")
for lang, idv, text, rec, verdict in res:
    if verdict == "候補": print(f"  [{lang}] {idv}: 本文[{text[:34]}]  ロケール認識[{rec[:34]}]")
