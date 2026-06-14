# 聞いて話せる英語 の会話/文法記録Excel(App Aと同方針)。学習対象=英語/母語=日本語の2列。
#  会話全言語_<date>.xlsx(テーマ毎シート・3レベル連結) / 文法全言語_<date>.xlsx(テーマ毎・可変問題数)。
import json, re, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "expo-app", "data")
SC = os.path.join(ROOT, "scripts")
DATE = sys.argv[1] if len(sys.argv) > 1 else "2026-06-14"
rd = lambda p: json.load(open(p, encoding="utf-8"))

themes = rd(os.path.join(DATA, "themes.json"))
gthemes = rd(os.path.join(DATA, "grammarThemes.json"))
conv = rd(os.path.join(SC, "_eng_conv.json"))      # {"T-L":[{en,ja}]}
gram = rd(os.path.join(SC, "_eng_grammar.json"))   # {"T":[{en,ja}]}

HDR = ["日本語", "英語"]
def sheet_name(prefix, num, name):
    return re.sub(r'[\\/?*\[\]:]', '・', f"{prefix}{int(num):02d}_{name}")[:31]

def build(out_path, themes_list, prefix, get_rows):
    wb = Workbook(); wb.remove(wb.active)
    bold = Font(bold=True)
    for t in sorted(themes_list, key=lambda x: x["id"]):
        ws = wb.create_sheet(sheet_name(prefix, t["id"], t["name"]))
        for c, h in enumerate(HDR, 1):
            ws.cell(row=1, column=c, value=h).font = bold
        r = 2
        for ja, en in get_rows(t["id"]):
            ws.cell(row=r, column=1, value=ja)
            ws.cell(row=r, column=2, value=en)
            r += 1
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 52
        for row in ws.iter_rows(min_row=2, max_row=r, max_col=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    wb.save(out_path)
    print(f"saved {os.path.basename(out_path)}: {len(wb.sheetnames)} sheets")

def conv_rows(tid):
    out = []
    for lv in (1, 2, 3):
        for it in conv.get(f"{tid}-{lv}", []):
            out.append((it.get("ja", ""), it.get("en", "")))
    return out

def gram_rows(tid):
    return [(it.get("ja", ""), it.get("en", "")) for it in gram.get(str(tid), [])]

build(os.path.join(ROOT, f"会話全言語_{DATE}.xlsx"), themes, "会話", conv_rows)
build(os.path.join(ROOT, f"文法全言語_{DATE}.xlsx"), gthemes, "文法", gram_rows)
